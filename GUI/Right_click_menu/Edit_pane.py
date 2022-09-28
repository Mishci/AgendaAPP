from tkinter import Toplevel, LabelFrame, Label, StringVar, PhotoImage, Button
from GUI.Menu.Edit_item.Entry_with_placeholder import EntryWithPlaceholder
from LOGIC.readDataFromZippedExcel import readDataFromZippedExcel
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from LOGIC.zip_file import zip_file


class Edit_Pane(Toplevel):
    def __init__(self, column_data_tuples_list, tree, iid, *args, **kwargs):
        Toplevel.__init__(self)
        self.geometry("630x520")
        self.config(bg="#7bacc7")
        self.title('  EDITUJ ZÁZNAM  ')
        self.iconphoto(True, PhotoImage(file="img.png"))
        self.attributes('-topmost', 'true')
        self.data_tuple_list = column_data_tuples_list
        self.tree = tree
        self.iid = iid

        self.check = print(f"self.tuple list = {column_data_tuples_list}")
        # TODO > join into single list
        self.personal_Svars = []
        self.datumy_stringVars = []
        self.cpp_stringVars = []

        # column name lists constants
        self.create_layout()

        # disabling the X-closing button
        self.grab_set()
        self.protocol("WM_DELETE_WINDOW", self.disable_close_on_X)

    def create_layout(self):
        # OSOBNI UDAJE LABELFRAME ----------------------------------------------------------------------
        Basic_data_LFrame = LabelFrame(self,
                                       text=f"Osobní údaje:                 {self.data_tuple_list[1][1]}  {self.data_tuple_list[0][1]}",
                                       relief="sunken", height=200, width=580,
                                       highlightcolor="LightSeaGreen", bg="lightblue", bd=3,
                                       highlightthickness=5, font=("Arial", 12, "bold"), foreground="red")
        Basic_data_LFrame.grid(row=0, column=0, columnspan=2, padx=20, pady=15)

        for i in range(0, 5, 2):
            user_input = StringVar()
            self.personal_Svars.append(user_input)
            # left label + entry
            Name_Label = Label(Basic_data_LFrame, text=self.data_tuple_list[i][0] + ":", bg="lightblue")
            Entry_name = EntryWithPlaceholder(Basic_data_LFrame, self.data_tuple_list[i][1], textvariable=user_input)
            Name_Label.grid(column=0, row=i, padx=(20, 10), pady=5)
            Entry_name.grid(column=1, row=i, padx=(20, 40), pady=5)

            # right label + entry
            user_input2 = StringVar()
            self.personal_Svars.append(user_input2)
            Surname_Label = Label(Basic_data_LFrame, text=self.data_tuple_list[i + 1][0] + ":", bg="lightblue")
            Entry_surname = EntryWithPlaceholder(Basic_data_LFrame, self.data_tuple_list[i + 1][1],
                                                 textvariable=user_input2)
            Surname_Label.grid(column=2, row=i, padx=20, pady=5)
            Entry_surname.grid(column=3, row=i, padx=(0, 20), pady=5)

        print(f"self.personalsvars = {self.personal_Svars}")
        # ///////////////////////////////////////////////////////////////////////////////////////////////////////////

        # DATUMY LABELFRAME ----------------------------------------------------------------------------------
        Datumy_data_LFrame = LabelFrame(self, text="Datumy", relief="sunken", height=300, width=200,
                                        highlightcolor="LightSeaGreen", bg="lightblue", bd=3, highlightthickness=5,
                                        font=("Arial", 12, "bold"), foreground="red")
        Datumy_data_LFrame.grid(row=1, column=0, columnspan=1, padx=(20, 5), pady=15, sticky="n")

        index = 0
        for item in self.data_tuple_list[6:9]:
            user_input = StringVar()
            self.datumy_stringVars.append(user_input)
            Name_Label = Label(Datumy_data_LFrame, text=item[0] + ":", bg="lightblue")
            Entry_name = EntryWithPlaceholder(Datumy_data_LFrame, item[1], textvariable=user_input)
            Name_Label.grid(column=0, row=index, padx=(20, 10), pady=5)
            Entry_name.grid(column=1, row=index, padx=(20, 40), pady=5)

            value = self.datumy_stringVars[index].get()
            print(f"value of {item[0]} is {value}")
            index += 1
        # /////////////////////////////////////////////////////////////////////////////////////////////////////////

        # CPP LABELFRAME ----------------------------------------------------------------------------------
        CP_data_LFrame = LabelFrame(self, text="Církevní příspěvky", relief="sunken", height=300, width=200,
                                    highlightcolor="LightSeaGreen", bg="lightblue", bd=3, highlightthickness=5,
                                    font=("Arial", 12, "bold"), foreground="red")
        CP_data_LFrame.grid(row=1, column=1, columnspan=1, pady=15, padx=(0, 20))

        index = 0
        for item in self.data_tuple_list[10:]:
            user_input = StringVar()
            self.cpp_stringVars.append(user_input)
            Name_Label = Label(CP_data_LFrame, text=item[0] + ":", bg="lightblue")
            Entry_name = EntryWithPlaceholder(CP_data_LFrame, item[1], textvariable=user_input)
            Name_Label.grid(column=0, row=index, padx=(20, 10), pady=5)
            Entry_name.grid(column=1, row=index, padx=(20, 40), pady=5)

            value = self.cpp_stringVars[index].get()
            print(f"value of {item[0]} is {value}")
            index += 1

            # JOIN ALL THE SVAR LISTS INTO SINGLE LIST: TODO> bad indexes /> better would be a dict
            svars_join = []
            for item in [self.personal_Svars, self.datumy_stringVars, self.cpp_stringVars]:
                if item is not None:
                    svars_join = svars_join + item
            print(f"svars_join > >> : {svars_join}")

        # SAVE BUTTON ----------------------------------------------------------------------------------------
        Save = Button(self, text=" Ulož změny",
                      command=lambda: self.saveedit(svars_join))  # TODO dopln spolecny Svarslit
        Save.grid(row=2, padx=(50, 50), pady=20, columnspan=2, sticky="EW")

    # ///////////////////////////////////////// Methods for updating and saving //////////////////////////////

    def saveedit(self, Svars_list):
        # 1. create the temporary excel file to read by openxl
        data = readDataFromZippedExcel()
        try:
            data.to_csv("TEMP/1.csv",
                        date_format='d%m%Y', index=False, header=True)
            csv = pd.read_csv("TEMP/1.csv")

            # converting csv to excel
            temp = pd.ExcelWriter("TEMP/data.xlsx")
            csv.to_excel(temp, index=None, header=True)
            temp.save()
        finally:
            os.remove("TEMP/1.csv")

        # 2. read the sheet in excel / searching for the cell + index of the given name + surname - > Tuple
        workbook = load_workbook("TEMP/data.xlsx")
        sheet = workbook.active

        # 3. iterate over every single col in the specified row containing the name + surname
        rowNum = self.searchValueInWorkbook(sheet, self.data_tuple_list[0][1], self.data_tuple_list[1][1])
        for row in sheet.iter_rows(min_row=rowNum, max_row=rowNum):
            index = 0
            for cell in row:
                if index < len(Svars_list) - 1 and cell.value != Svars_list[index] and Svars_list[
                    index] is not None:
                    sheet[cell.coordinate] = Svars_list[index].get()
                index += 1

        workbook.save(filename="TEMP/data.xlsx")

        # 4. update the Zip file

        os.remove("PERSONAL/Hovorany/excel_back_Hovorany.zip")
        data = pd.read_excel("TEMP/data.xlsx")
        data["Nar."] = pd.to_datetime(data["Nar."], errors='coerce')  # timestamp get rid
        data["Nar."] = data["Nar."].dt.date  # overwrite |Nar.] values with no timestamp
        data.to_excel(f"PERSONAL/Hovorany/excel_back_Hovorany.xlsx", index=None, header=True)

        zip_file("PERSONAL/Hovorany/excel_back_Hovorany.xlsx", "a")

        # update the treeview and delete Temporary files
        self.tree.item (self.iid, values = tuple ([ val.get() for array in [self.personal_Svars, self.datumy_stringVars, self.cpp_stringVars]  for val in array]))
        self.destroy()

        temp.close() # TODO > Calling close() on already closed file.
        os.remove("TEMP/data.xlsx")
        os.remove("PERSONAL/Hovorany/excel_back_Hovorany.xlsx")

    def searchValueInWorkbook(self, ws, surname_string, name_string):
        row_index = 1
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row + 1):
            column = 1
            for col in row:
                coordinate = "{}{}".format(get_column_letter(column), row_index)
                if ws[coordinate].value == surname_string:
                    coordinate_plus1 = "{}{}".format(get_column_letter(column + 1), row_index)
                    if ws[coordinate_plus1].value == name_string:
                        return row_index
                column += 1
            row_index += 1
        return None

    def disable_close_on_X(self):
        self.destroy()
