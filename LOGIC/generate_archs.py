import os.path
from tkinter import LabelFrame, Entry, Button, StringVar, RAISED, CENTER, messagebox
import pandas as pd
import openpyxl
from openpyxl.styles import Side, Border, Alignment, Font
from datetime import datetime

from LOGIC.readDataFromZippedExcel import readDataFromZippedExcel


# ---------------------- crate form for adding "poverena osoba" ----------------------------------------------------
def pop_up_gui(adress, mainframe):
    widgetlist = mainframe.winfo_children ()
    if len(widgetlist) > 0:
        for widgets in mainframe.winfo_children():
            widgets.destroy()

    pop_up_frame = LabelFrame(mainframe, labelanchor="n", text=f"Pověřená osoba pro \n \n {adress}",
                              relief=RAISED, height=50, width=100, padx=50, pady=20, bg="lightblue")
    pop_up_frame.config(font=("Arial", 9))
    pop_up_frame.pack()
    poverena_osoba = StringVar(pop_up_frame)

    pop_up_entry = Entry(master=pop_up_frame, textvariable=poverena_osoba, justify=CENTER, width=50)
    pop_up_entry.pack(pady=20)
    pop_up_entry.focus_set()

    ok_button = Button(master=pop_up_frame, text="Potvrdit",
                       command=lambda: gen_styled_archs(adress, poverena_osoba.get()))
    ok_button.pack()

    return poverena_osoba


# ------------------------------------ CREATE ARCH FILE --------------------------------------------------------------
def generuj_archy(adresa):
    adress = adresa

    data = readDataFromZippedExcel()
    mask = data["Adresa"].str.contains(f"{adress}", na=False)  # column mask with condition
    columnList = data.columns.values
    CPPcols = [col for col in columnList if col.startswith('CPP')]

    # getting the adults filter

    eighteenPlus = []
    for item in list((data["Nar."].fillna(0))):
        if item == 0:
            pass
        elif item != 0:
            itemarray = str(item).split("-")
            print(int(itemarray[0]))
            if int(itemarray[0]) < (datetime.now().year - 18):
                eighteenPlus.append(item)

    # applying the mask condition
    data = data.loc[mask, ["Přijmení", "Jméno", "Adresa", "Město", "Nar.", CPPcols[-1]]]

    # applying the +18 condition
    arch = pd.DataFrame(
        data.loc[data["Nar."].isin(eighteenPlus), ["Přijmení", "Jméno", "Adresa", "Město",
                                                   CPPcols[-1]]])  # with the last CPP column
    arch = arch.reset_index(drop=True)  # resetting index that it would start with 1
    arch.index += 1
    arch = arch.style.set_properties(**{"text-align": "center", "border": "2px"})

    # create excel file based on the mask
    try:
        writer = pd.ExcelWriter(f"Vyberci_Arch_{adress}.xlsx", engine='openpyxl')
        arch.to_excel(writer, sheet_name=f"{adress}", startrow=5, startcol=1)
        writer.save()
        writer.close()
    except PermissionError:
        messagebox.showerror(title="SOUBOR JE POUŽÍVÁN JINýM PROGRAMEM",
                             message=f"Zavřete soubor agenda/Vyberci_Arch_{adress}.xlsx "
                                     f"ve všech programech a vygenerujte arch znovu! \n "
                                     f"Tímto krokem aktualizujete obsah archu")


# --------------------------------- STYLING TABULKY ----------------------------------------------------------
def borders(adress, osoba):
    try:
        if os.path.exists(f"C:/Users/Michal/PycharmProjects/agenda/Vyberci_Arch_{adress}.xlsx"):
            adresa = adress
            person = osoba

            wb = openpyxl.load_workbook(f"C:/Users/Michal/PycharmProjects/agenda/Vyberci_Arch_{adress}.xlsx",
                                        data_only=True)
            # ---------------Borders round the cells --------------------------------
            ws = wb.active
            thin = Side(border_style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            print("ok")
            ws.column_dimensions["A"].width = 5
            ws.column_dimensions["B"].width = 5
            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 15
            ws.column_dimensions["E"].width = 20
            ws.column_dimensions["F"].width = 10

            for i in range(6, int(ws.max_row)):
                row = ws.row_dimensions[i + 1]
                row.height = 25

            for row in ws.iter_rows(min_row=7, min_col=2):
                for cell in row:
                    cell.border = border

            # ------------------------ Hlavicka merging + styling ---------------------------------------------
            # Nadpis
            from LOGIC.login_form_entries_validation import userGlobal
            ws.merge_cells('B1:G1')
            nadpis = ws.cell(row=1, column=2)
            nadpis.value = f"Výběrčí arch NO CČSH v {userGlobal}"
            nadpis.alignment = Alignment(horizontal='center', vertical='center')
            nadpis.font = Font(size=14, bold=True)

            # rok
            ws.merge_cells('B2:G2')
            rok = ws.cell(row=2, column=2)
            rok.value = f"pro rok {datetime.now().year} "
            rok.alignment = Alignment(horizontal='center', vertical='center')
            rok.font = Font(size=14, bold=True)

            # Poverena osoba
            ws.merge_cells('A4:D4')
            osoba = ws['A4']
            osoba.value = f"pověřená osoba: {person}"
            osoba.alignment = Alignment(horizontal='center', vertical='center')
            osoba.font = Font(size=12, bold=False)

            # Misto
            ws.merge_cells('E4:G4')
            misto = ws['E4']
            misto.value = f"místo: {adress}"
            misto.alignment = Alignment(horizontal='center', vertical='center')
            misto.font = Font(size=12, bold=False)

            # -------------------------- saving ----------------------------------------------------
            wb.save(f"Vyberci_Arch_{adresa}.xlsx")

    except FileNotFoundError:
        messagebox.showerror(title="EXCEL SOUBOR NENALEZEN", message="Soubor byl smazán nebo přesunut!")


# ----------------------------souhrnna funkce pro menubutton triggering -----------------------------------------------
def gen_styled_archs(adresa, osoba):
    generuj_archy(adresa)
    borders(adresa, osoba)
